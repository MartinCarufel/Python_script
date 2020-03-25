import openpyxl
import tkinter.filedialog
import csv

# source_path = tkinter.filedialog.askopenfilename()
# target_path = tkinter.filedialog.asksaveasfilename()

file_name = "detect_file_list.csv"
path_source = "C:/Drive-W/test_veh_profile/detect2/Excel/from_old_temp/"
path_target = "C:/Drive-W/test_veh_profile/detect2/Excel/"


def copy_sheet_basic_sensings(source_path, target_path,  filename):
    origin = openpyxl.load_workbook(filename=source_path + filename, read_only=False, keep_vba=True)
    target = openpyxl.load_workbook(filename=target_path + filename, read_only=False, keep_vba=True)
    origin_sheet = origin["Basic Sensings"]
    target_sheet = target["Basic Sensings"]

    for i in range(4, 100):
        for j in range(1, 19):
            target_sheet.cell(row=i,column=j).value = origin_sheet.cell(row=i,column=j).value
    target.save(target_path + filename)


def copy_sheet_function(source_path, target_path,  filename):
    origin = openpyxl.load_workbook(filename=source_path + filename, read_only=False, keep_vba=True)
    target = openpyxl.load_workbook(filename=target_path + filename, read_only=False, keep_vba=True)
    origin_sheet = origin["Functions"]
    target_sheet = target["Functions"]

    for i in range(5, 100):
        for j in range(1, 15):
            target_sheet.cell(row=i,column=j).value = origin_sheet.cell(row=i,column=j).value
    target.save(target_path + filename)


with open(file_name, 'r') as csvfile:
    csvreader = csv.reader(csvfile)
    for line in csvreader:
        for item in range(len(line)):
            copy_sheet_basic_sensings(path_source, path_target, line[item])
            print(line[item])

